"""
Generate a market-research workbook for iPad cases (pochettes iPad) targeted
at the Moroccan market.

Output: docs/research/pochettes-ipad-2010-2026.xlsx

Five sheets:
  1. README              — methodology + how to use
  2. Top Stocking        — purchase priority across all price tiers
  3. iPad Models         — full reference 2010-2026
  4. Cases Catalog       — recommended case SKUs by iPad model
  5. Sources             — links

Data is calibrated against real Jumia.ma prices (April 2026 snapshot)
and cross-referenced with Amazon / Engadget / Macworld bestseller lists.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


ROOT = Path(__file__).resolve().parents[3]
OUT = ROOT / "docs" / "research" / "pochettes-ipad-2010-2026.xlsx"

# ── Palette ─────────────────────────────────────────────────────────────
BLUE_DARK   = "1F4E79"
BLUE_LIGHT  = "D9E2F3"
GREEN_LIGHT = "E2EFDA"
GREEN_DARK  = "548235"
AMBER_LIGHT = "FFF2CC"
RED_LIGHT   = "F8CBAD"
GREY_LIGHT  = "F2F2F2"
GREY_DARK   = "595959"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

H_FILL = PatternFill("solid", start_color=BLUE_DARK, end_color=BLUE_DARK)
H_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws: Worksheet, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.alignment = H_ALIGN
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def autosize(ws: Worksheet, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def freeze_header(ws: Worksheet, row: int = 2) -> None:
    ws.freeze_panes = ws.cell(row=row, column=1)


def write_title(ws: Worksheet, text: str, row: int, span: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", size=16, bold=True, color=BLUE_DARK)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 26


# ────────────────────────────────────────────────────────────────────────
# Sheet 1: README
# ────────────────────────────────────────────────────────────────────────
def build_readme(ws: Worksheet) -> None:
    ws.title = "README"
    write_title(ws, "Pochettes iPad — Étude de stockage marché marocain", 1, 4)

    rows = [
        ("Date de l'étude", "1 mai 2026"),
        ("Marché cible", "Maroc — boutique tech (teliphoni)"),
        ("Devise", "MAD (Dirham marocain) — HT"),
        ("Périmètre iPad", "Toutes générations 2010 → 2026 (44 modèles distincts)"),
        ("Positionnement prix", "Toutes gammes (entrée 30 MAD → premium 1 500 MAD)"),
        ("", ""),
        ("Méthodologie de classement", ""),
        ("  · Popularité du modèle iPad", "Volumes installés au Maroc + ventes en cours"),
        ("  · Disponibilité fournisseur", "Jumia.ma + AliExpress + import Amazon FR"),
        ("  · Marge pratiquée", "Estimation 30-60 % selon segment"),
        ("  · Saturation marché", "Concurrence Jumia / Marjane / souks"),
        ("", ""),
        ("Comment utiliser ce fichier", ""),
        ("  Onglet Top Stocking", "Liste à acheter par ordre de priorité 1→5"),
        ("  Onglet iPad Models", "Référentiel pour identifier compatibilité d'une pochette"),
        ("  Onglet Cases Catalog", "Détail par modèle iPad : 3-5 SKUs recommandés"),
        ("  Onglet Sources", "Liens et URLs consultés"),
        ("", ""),
        ("Légende priorité", ""),
        ("  P1 — Must-stock", "Volume max — à avoir absolument en stock"),
        ("  P2 — Forte demande", "Complément — stocker en quantité raisonnable"),
        ("  P3 — Demande modérée", "Stock minimal pour parc installé"),
        ("  P4 — Anticipation 2026", "Modèles à venir — précommander chez fournisseur"),
        ("  P5 — Niche / collector", "À la commande seulement — pas de stock"),
        ("", ""),
        ("Hypothèses prix (HT MAD)", ""),
        ("  Entrée de gamme", "30 — 150 MAD"),
        ("  Milieu de gamme", "150 — 400 MAD"),
        ("  Premium", "400 — 1 500 MAD"),
        ("", ""),
        ("AVERTISSEMENT", "Les prix sont indicatifs — vérifier avec fournisseur avant achat. Les modèles 2026 (iPad Air M4, iPad 12, Mini 8 OLED) sont basés sur rumeurs Apple — caractéristiques à confirmer à la sortie officielle."),
    ]

    for i, (label, val) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(name="Calibri", size=11, bold=bool(label and not label.startswith(" ")))
        ws.cell(row=i, column=2, value=val).font = Font(name="Calibri", size=11)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        if label and not label.startswith(" ") and label not in ("AVERTISSEMENT", ""):
            ws.cell(row=i, column=1).fill = PatternFill("solid", start_color=BLUE_LIGHT, end_color=BLUE_LIGHT)
        if label == "AVERTISSEMENT":
            ws.cell(row=i, column=1).fill = PatternFill("solid", start_color=AMBER_LIGHT, end_color=AMBER_LIGHT)
            ws.cell(row=i, column=2).fill = PatternFill("solid", start_color=AMBER_LIGHT, end_color=AMBER_LIGHT)
            ws.row_dimensions[i].height = 50

    autosize(ws, {"A": 30, "B": 80})


# ────────────────────────────────────────────────────────────────────────
# Sheet 2: Top Stocking Priority
# ────────────────────────────────────────────────────────────────────────
TOP_STOCKING = [
    # (priority, ipad_target, case_type, brand, segment, price_min, price_max, justification)
    # ── P1 must-stock ─────────────────────────────────────────────────
    (1, "iPad 11 (A16, 2025) 11\"",        "Book Cover 360° + crayon",   "Generic / Case",     "Milieu",  150, 250, "Modèle iPad le plus vendu actuellement — boost A19 attendu mais 2025 reste leader"),
    (1, "iPad 10 (2022) 10.9\"",            "Folio Smart Cover",          "Generic / Case Book", "Milieu",  150, 250, "Énorme parc installé, USB-C, encore activement vendu"),
    (1, "iPad 9 (2021) 10.2\"",             "Folio basique magnétique",   "Itel / Generic",     "Entrée",   60, 150, "Dernier iPad Lightning, parc gigantesque école+famille"),
    (1, "iPad Air 11\" M2 (2024)",          "Smart Folio + porte Pencil", "Generic / ESR",      "Milieu",  200, 350, "Volume mid-range, recherche élevée, marge correcte"),
    (1, "iPad Mini 7 (2024) 8.3\"",         "Folio compact",              "Generic / Fintie",   "Milieu",  150, 280, "Mini = niche fidèle, peu de concurrence locale"),
    (1, "Universel 10.2\" / 10.9\"",        "Pochette néoprène / sleeve", "Itel",               "Entrée",   29,  79, "Vente d'impulsion, marge très élevée, multi-modèles"),
    # ── P2 forte demande ──────────────────────────────────────────────
    (2, "iPad Pro 11\" M4 (2024)",          "Smart Folio premium",        "ESR / Apple",        "Premium", 400, 900, "Clientèle pro, prix élevé = marge unitaire"),
    (2, "iPad Air 13\" M2 (2024)",          "Folio grand format",         "Generic / ESR",      "Milieu",  250, 450, "Format 13\" en croissance"),
    (2, "iPad Mini 6 (2021)",               "Folio compact 8.3\"",        "Generic / MoKo",     "Milieu",  120, 220, "Parc installé Mini 6 actif"),
    (2, "iPad Air 5 (2022) 10.9\"",         "Smart Folio",                "Generic / Case",     "Milieu",  150, 280, "Compatible aussi Air 4 — gain de stock"),
    (2, "iPad Pro 13\" M4 (2024)",          "Folio premium",              "ESR / Apple",        "Premium", 500, 1200, "Niche pro premium, demande de qualité"),
    (2, "iPad Air M3 11\" (2025)",          "Smart Folio compatible M2",  "Generic",            "Milieu",  180, 320, "Compatible mécaniquement avec étuis Air M2 (même dimensions)"),
    # ── P3 demande modérée ────────────────────────────────────────────
    (3, "iPad Air 4 (2020) 10.9\"",         "Folio générique",            "Generic",            "Entrée",  100, 200, "Encore en circulation, occasion forte"),
    (3, "iPad Pro 11\" 2018/2020/2021/2022","Folio compatible toutes gen","Generic",            "Milieu",  150, 300, "Une seule pochette pour 4 générations — gain stock énorme"),
    (3, "iPad 7 / iPad 8 (10.2\")",         "Folio basique",              "Itel / Generic",     "Entrée",   59, 130, "Compatible avec iPad 9 — mutualiser le stock"),
    (3, "iPad Pro 12.9\" 2018-2022",        "Folio grand format",         "Generic",            "Milieu",  200, 400, "Compatible 4 gens — bon ratio stock/couverture"),
    (3, "iPad Mini 5 (2019)",               "Folio compact 7.9\"",        "Generic",            "Entrée",   60, 150, "Demande résiduelle"),
    (3, "Universel 11\"-13\"",              "Pochette à fermeture éclair","Itel / Generic",     "Entrée",   49,  99, "Sleeve universelle — compatible majorité iPad récents"),
    # ── P4 anticipation 2026 ──────────────────────────────────────────
    (4, "iPad Air M4 (2026)",               "Smart Folio (compat. M2/M3)","Generic / ESR",      "Milieu",  200, 350, "Sortie mars 2026 — design identique M2/M3, stock recyclable"),
    (4, "iPad 12 (A19, 2026)",              "Folio Smart Cover",          "Generic",            "Entrée",  120, 230, "Sortie mid/late 2026 — anticiper avec fournisseur"),
    (4, "iPad Mini 8 OLED (2026)",          "Folio premium compact",      "ESR / Apple",        "Premium", 300, 600, "OLED + A18 Pro — clientèle premium attendue"),
    # ── P5 niche / collector ──────────────────────────────────────────
    (5, "iPad Pro 9.7\" (2016)",            "Folio sur commande",         "Generic",            "Entrée",   80, 180, "Très peu de demande — éviter stock"),
    (5, "iPad Air 2 (2014)",                "Folio sur commande",         "Generic",            "Entrée",   60, 150, "Collector / réparation"),
    (5, "iPad 5/6 (2017/2018) 9.7\"",       "Folio sur commande",         "Generic",            "Entrée",   60, 130, "Demande résiduelle"),
]


def build_top_stocking(ws: Worksheet) -> None:
    headers = [
        "Priorité", "iPad cible", "Type de pochette", "Marque(s)",
        "Segment prix", "Prix min (MAD HT)", "Prix max (MAD HT)",
        "Prix moy. (MAD HT)", "Justification",
    ]
    write_title(ws, "Top Stocking — Priorité d'achat", 1, len(headers))

    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(headers))

    for i, row in enumerate(TOP_STOCKING, start=4):
        prio, target, ctype, brand, seg, p_min, p_max, just = row
        ws.cell(row=i, column=1, value=f"P{prio}")
        ws.cell(row=i, column=2, value=target)
        ws.cell(row=i, column=3, value=ctype)
        ws.cell(row=i, column=4, value=brand)
        ws.cell(row=i, column=5, value=seg)
        ws.cell(row=i, column=6, value=p_min)
        ws.cell(row=i, column=7, value=p_max)
        # Average — Excel formula so it stays dynamic
        ws.cell(row=i, column=8, value=f"=AVERAGE(F{i}:G{i})")
        ws.cell(row=i, column=9, value=just)

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(c == 9),
                horizontal=("center" if c in (1, 5, 6, 7, 8) else "left"),
            )
            if c in (6, 7, 8):
                cell.number_format = "#,##0 \"MAD\""

        ws.row_dimensions[i].height = 36

    # Conditional formatting on Priority column
    last_row = 3 + len(TOP_STOCKING)
    color_map = {
        "P1": ("16A34A", "FFFFFF"),  # green
        "P2": ("0EA5E9", "FFFFFF"),  # blue
        "P3": ("F59E0B", "111111"),  # amber
        "P4": ("8B5CF6", "FFFFFF"),  # purple
        "P5": ("9CA3AF", "111111"),  # grey
    }
    for label, (bg, fg) in color_map.items():
        rule = Rule(
            type="cellIs", operator="equal",
            formula=[f'"{label}"'],
            dxf=DifferentialStyle(
                fill=PatternFill("solid", start_color=bg, end_color=bg),
                font=Font(name="Calibri", size=11, bold=True, color=fg),
            ),
        )
        ws.conditional_formatting.add(f"A4:A{last_row}", rule)

    # Color-band by Segment column
    seg_colors = {
        "Entrée":  GREEN_LIGHT,
        "Milieu":  BLUE_LIGHT,
        "Premium": AMBER_LIGHT,
    }
    for label, bg in seg_colors.items():
        rule = Rule(
            type="cellIs", operator="equal",
            formula=[f'"{label}"'],
            dxf=DifferentialStyle(fill=PatternFill("solid", start_color=bg, end_color=bg)),
        )
        ws.conditional_formatting.add(f"E4:E{last_row}", rule)

    autosize(ws, {"A": 10, "B": 32, "C": 32, "D": 22, "E": 12, "F": 16, "G": 16, "H": 16, "I": 60})
    freeze_header(ws, row=4)


# ────────────────────────────────────────────────────────────────────────
# Sheet 3: iPad Models 2010-2026
# ────────────────────────────────────────────────────────────────────────
IPAD_MODELS = [
    # (id, name, year, family, screen, connector, pencil, status, notes)
    (1,  "iPad (1ʳᵉ gen)",                 2010, "iPad",       "9.7\" LCD",       "30-pin",   "—",                    "Obsolète",   "Lancement de la gamme"),
    (2,  "iPad 2",                         2011, "iPad",       "9.7\" LCD",       "30-pin",   "—",                    "Obsolète",   "—"),
    (3,  "iPad (3ᵉ gen)",                  2012, "iPad",       "9.7\" Retina",    "30-pin",   "—",                    "Obsolète",   "1ᵉʳ Retina"),
    (4,  "iPad (4ᵉ gen)",                  2012, "iPad",       "9.7\" Retina",    "Lightning","—",                    "Obsolète",   "1ᵉʳ Lightning"),
    (5,  "iPad Mini",                      2012, "iPad Mini",  "7.9\" LCD",       "Lightning","—",                    "Obsolète",   "—"),
    (6,  "iPad Mini 2",                    2013, "iPad Mini",  "7.9\" Retina",    "Lightning","—",                    "Obsolète",   "—"),
    (7,  "iPad Air",                       2013, "iPad Air",   "9.7\" Retina",    "Lightning","—",                    "Obsolète",   "1ᵉʳ Air, plus fin"),
    (8,  "iPad Mini 3",                    2014, "iPad Mini",  "7.9\" Retina",    "Lightning","—",                    "Obsolète",   "Touch ID"),
    (9,  "iPad Air 2",                     2014, "iPad Air",   "9.7\" Retina",    "Lightning","—",                    "Rare",       "Demande de remplacement résiduelle"),
    (10, "iPad Mini 4",                    2015, "iPad Mini",  "7.9\" Retina",    "Lightning","—",                    "Obsolète",   "—"),
    (11, "iPad Pro 12.9\" (1ʳᵉ gen)",     2015, "iPad Pro",   "12.9\" LCD",      "Lightning","Pencil 1",             "Obsolète",   "1ᵉʳ Pro grand format"),
    (12, "iPad Pro 9.7\"",                 2016, "iPad Pro",   "9.7\" LCD",       "Lightning","Pencil 1",             "Rare",       "True Tone"),
    (13, "iPad (5ᵉ gen)",                  2017, "iPad",       "9.7\" LCD",       "Lightning","Pencil 1",             "Obsolète",   "—"),
    (14, "iPad Pro 12.9\" (2ᵉ gen)",      2017, "iPad Pro",   "12.9\" LCD",      "Lightning","Pencil 1",             "Obsolète",   "ProMotion 120 Hz"),
    (15, "iPad Pro 10.5\"",                2017, "iPad Pro",   "10.5\" LCD",      "Lightning","Pencil 1",             "Rare",       "Format intermédiaire"),
    (16, "iPad (6ᵉ gen)",                  2018, "iPad",       "9.7\" LCD",       "Lightning","Pencil 1",             "Rare",       "—"),
    (17, "iPad Pro 11\" (1ʳᵉ gen)",       2018, "iPad Pro",   "11\" LCD",        "USB-C",    "Pencil 2",             "Occasion",   "1ᵉʳ Pro Face ID + USB-C"),
    (18, "iPad Pro 12.9\" (3ᵉ gen)",      2018, "iPad Pro",   "12.9\" LCD",      "USB-C",    "Pencil 2",             "Occasion",   "—"),
    (19, "iPad Air (3ᵉ gen)",              2019, "iPad Air",   "10.5\" LCD",      "Lightning","Pencil 1",             "Rare",       "Touch ID home"),
    (20, "iPad Mini 5",                    2019, "iPad Mini",  "7.9\" Retina",    "Lightning","Pencil 1",             "Rare",       "Dernier Mini Lightning"),
    (21, "iPad (7ᵉ gen)",                  2019, "iPad",       "10.2\" LCD",      "Lightning","Pencil 1",             "Vendu",      "Format 10.2\" introduit"),
    (22, "iPad Pro 11\" (2ᵉ gen)",        2020, "iPad Pro",   "11\" LCD",        "USB-C",    "Pencil 2",             "Vendu",      "—"),
    (23, "iPad Pro 12.9\" (4ᵉ gen)",      2020, "iPad Pro",   "12.9\" LCD",      "USB-C",    "Pencil 2",             "Vendu",      "—"),
    (24, "iPad Air (4ᵉ gen)",              2020, "iPad Air",   "10.9\" LCD",      "USB-C",    "Pencil 2",             "Vendu",      "1ᵉʳ Air USB-C, design Pro"),
    (25, "iPad (8ᵉ gen)",                  2020, "iPad",       "10.2\" LCD",      "Lightning","Pencil 1",             "Vendu",      "—"),
    (26, "iPad Pro 11\" (3ᵉ gen) M1",     2021, "iPad Pro",   "11\" LCD M1",     "USB-C",    "Pencil 2",             "Vendu",      "Apple Silicon"),
    (27, "iPad Pro 12.9\" (5ᵉ gen) M1",   2021, "iPad Pro",   "12.9\" mLED M1",  "USB-C",    "Pencil 2",             "Vendu",      "Mini-LED XDR"),
    (28, "iPad (9ᵉ gen) ★",                2021, "iPad",       "10.2\" LCD",      "Lightning","Pencil 1",             "Vendu (top)","TOP volumes — dernier Lightning"),
    (29, "iPad Mini 6 ★",                  2021, "iPad Mini",  "8.3\" LCD",       "USB-C",    "Pencil 2",             "Vendu (top)","Refonte Mini USB-C"),
    (30, "iPad Air (5ᵉ gen) M1 ★",        2022, "iPad Air",   "10.9\" LCD M1",   "USB-C",    "Pencil 2",             "Vendu (top)","M1 + 5G"),
    (31, "iPad Pro 11\" (4ᵉ gen) M2",     2022, "iPad Pro",   "11\" LCD M2",     "USB-C",    "Pencil 2 + Hover",     "Vendu",      "Pencil hover"),
    (32, "iPad Pro 12.9\" (6ᵉ gen) M2",   2022, "iPad Pro",   "12.9\" mLED M2",  "USB-C",    "Pencil 2 + Hover",     "Vendu",      "—"),
    (33, "iPad (10ᵉ gen) ★",               2022, "iPad",       "10.9\" LCD",      "USB-C",    "Pencil 1 (USB-C adpt)","Vendu (top)","TOP volume USB-C entrée"),
    (34, "iPad Air 11\" M2 ★",             2024, "iPad Air",   "11\" LCD M2",     "USB-C",    "Pencil Pro / USB-C",   "Vendu (top)","Air 11\" actuel"),
    (35, "iPad Air 13\" M2",               2024, "iPad Air",   "13\" LCD M2",     "USB-C",    "Pencil Pro / USB-C",   "Vendu",      "1ᵉʳ Air 13\""),
    (36, "iPad Pro 11\" M4 ★",             2024, "iPad Pro",   "11\" OLED M4",    "USB-C",    "Pencil Pro / USB-C",   "Vendu (top)","OLED Tandem"),
    (37, "iPad Pro 13\" M4",               2024, "iPad Pro",   "13\" OLED M4",    "USB-C",    "Pencil Pro / USB-C",   "Vendu",      "OLED Tandem grand format"),
    (38, "iPad Mini 7 (A17 Pro) ★",        2024, "iPad Mini",  "8.3\" LCD",       "USB-C",    "Pencil Pro / USB-C",   "Vendu (top)","Apple Intelligence"),
    (39, "iPad (11ᵉ gen) — A16 ★",         2025, "iPad",       "11\" LCD A16",    "USB-C",    "Pencil USB-C",         "Vendu (top)","TOP volume actuel"),
    (40, "iPad Air M3 11\" / 13\"",        2025, "iPad Air",   "11\"/13\" LCD M3","USB-C",    "Pencil Pro / USB-C",   "Vendu",      "Compat. mécanique M2"),
    (41, "iPad Pro M5 11\" / 13\" ★",      2025, "iPad Pro",   "OLED M5",         "USB-C / Thunder","Pencil Pro",     "Vendu (top)","Apple Intelligence Pro"),
    (42, "iPad Air M4 11\" / 13\"",        2026, "iPad Air",   "11\"/13\" LCD M4","USB-C",    "Pencil Pro / USB-C",   "À venir",    "Mars 2026 — design inchangé"),
    (43, "iPad (12ᵉ gen) — A19",           2026, "iPad",       "11\" LCD A19",    "USB-C",    "Pencil USB-C",         "À venir",    "Mid/late 2026 — Apple Intelligence"),
    (44, "iPad Mini 8 OLED (A18 Pro)",     2026, "iPad Mini",  "8.3\" OLED",      "USB-C",    "Pencil Pro / USB-C",   "À venir",    "1ᵉʳ Mini OLED — fin 2026"),
]


def build_ipad_models(ws: Worksheet) -> None:
    headers = ["#", "Modèle", "Année", "Famille", "Écran", "Connecteur",
               "Apple Pencil", "Statut", "Notes"]
    write_title(ws, "Modèles iPad — Référentiel 2010 → 2026", 1, len(headers))

    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(headers))

    for i, m in enumerate(IPAD_MODELS, start=4):
        for c, val in enumerate(m, start=1):
            ws.cell(row=i, column=c, value=val)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center",
                horizontal=("center" if c in (1, 3, 4, 6, 8) else "left"),
                wrap_text=(c == 9),
            )
            cell.font = Font(name="Calibri", size=10)
        ws.row_dimensions[i].height = 22

    # Status color coding
    last = 3 + len(IPAD_MODELS)
    status_colors = {
        "Vendu (top)": ("16A34A", "FFFFFF"),
        "Vendu":       (GREEN_LIGHT, "111111"),
        "À venir":     ("8B5CF6", "FFFFFF"),
        "Occasion":    (BLUE_LIGHT, "111111"),
        "Rare":        (AMBER_LIGHT, "111111"),
        "Obsolète":    ("E5E7EB", "6B7280"),
    }
    for label, (bg, fg) in status_colors.items():
        rule = Rule(
            type="cellIs", operator="equal",
            formula=[f'"{label}"'],
            dxf=DifferentialStyle(
                fill=PatternFill("solid", start_color=bg, end_color=bg),
                font=Font(name="Calibri", size=10, bold=True, color=fg),
            ),
        )
        ws.conditional_formatting.add(f"H4:H{last}", rule)

    autosize(ws, {"A": 5, "B": 30, "C": 8, "D": 12, "E": 22, "F": 16, "G": 22, "H": 14, "I": 36})
    freeze_header(ws, row=4)


# ────────────────────────────────────────────────────────────────────────
# Sheet 4: Cases Catalog by iPad model
# ────────────────────────────────────────────────────────────────────────
CASES_CATALOG = [
    # (ipad_target, sku_label, type, brand_segment, segment, price_min, price_max, source_indicator)
    # iPad 11 (A16, 2025) 11" — top
    ("iPad 11 (A16) 11\"",  "Case Book Cover 360°",          "Folio rotatif",   "Generic / Case",        "Milieu",  95, 199, "Jumia ★★"),
    ("iPad 11 (A16) 11\"",  "Generic Pochette Cover Antichoc","Folio anti-choc","Generic",               "Milieu", 199, 349, "Jumia ★"),
    ("iPad 11 (A16) 11\"",  "Generic Cover 360 Rotation",    "Rotatif 360°",    "Generic",               "Milieu",  99, 199, "Jumia"),
    ("iPad 11 (A16) 11\"",  "Apple Smart Folio iPad 11",     "Smart Folio",     "Apple",                 "Premium", 850, 1100, "Apple/Bestmark"),
    ("iPad 11 (A16) 11\"",  "ESR Rebound Magnetic",          "Folio Pencil slot","ESR (importé)",        "Premium", 350, 550, "Amazon FR"),
    # iPad 10 (2022) 10.9"
    ("iPad 10 (2022) 10.9\"", "Case Book Cover iPad 10",     "Folio basique",   "Generic / Case Book",   "Milieu", 199, 249, "Jumia ★★"),
    ("iPad 10 (2022) 10.9\"", "Pochette Antichoc PU+TPU",    "Folio anti-choc", "Generic",               "Milieu", 159, 280, "Jumia"),
    ("iPad 10 (2022) 10.9\"", "Apple Smart Folio (iPad 10)", "Smart Folio",     "Apple",                 "Premium", 750, 950, "Apple/Bestmark"),
    ("iPad 10 (2022) 10.9\"", "Logitech Combo Touch",        "Folio + clavier", "Logitech",              "Premium",1500, 2200, "Bestmark/Import"),
    # iPad 9 (2021) 10.2"
    ("iPad 9 (2021) 10.2\"",  "Itel Pochette Tablette PAD 2","Pochette néoprène","Itel",                 "Entrée",   29,  69, "Jumia ★★★ (top sales)"),
    ("iPad 9 (2021) 10.2\"",  "Generic Folio 10.2\"",        "Folio basique",   "Generic",               "Entrée",   59, 130, "Jumia/Avito"),
    ("iPad 9 (2021) 10.2\"",  "Smart Cover compatible 9/8/7","Folio multi-gen", "Generic",               "Milieu",  120, 220, "Jumia"),
    ("iPad 9 (2021) 10.2\"",  "ESR Rebound Pencil Holder",   "Folio Pencil",    "ESR",                   "Milieu",  220, 380, "Amazon FR"),
    ("iPad 9 (2021) 10.2\"",  "Logitech Crayon + Combo Touch","Pack productivité","Logitech",            "Premium",1400, 2000, "Import"),
    # iPad Air 11" M2 (2024)
    ("iPad Air 11\" M2 (2024)","Generic Smart Folio Air 11", "Smart Folio",     "Generic",               "Milieu",  180, 320, "Jumia"),
    ("iPad Air 11\" M2 (2024)","Pochette Antichoc Cuir PU",  "Folio cuir",      "Generic",               "Milieu",  220, 380, "Jumia ★"),
    ("iPad Air 11\" M2 (2024)","Apple Smart Folio Air 11",   "Smart Folio",     "Apple",                 "Premium", 800, 1100, "Apple"),
    ("iPad Air 11\" M2 (2024)","ESR Rebound Magnetic Air",   "Folio premium",   "ESR",                   "Premium", 380, 600, "Amazon FR"),
    # iPad Mini 7 (2024) 8.3"
    ("iPad Mini 7 (2024)",    "Folio Mini 8.3\" compatible 6","Folio compact",  "Generic",               "Milieu",  120, 220, "Jumia"),
    ("iPad Mini 7 (2024)",    "ESR Rebound Mini",            "Folio premium",   "ESR",                   "Milieu",  280, 450, "Amazon FR"),
    ("iPad Mini 7 (2024)",    "Apple Smart Folio Mini",      "Smart Folio",     "Apple",                 "Premium", 700, 900, "Apple"),
    # Universel pochette
    ("Universel 10.2\"-10.9\"","Sleeve néoprène universelle","Pochette zip",    "Itel / Generic",        "Entrée",   29,  79, "Jumia ★★ (impulsion)"),
    ("Universel 11\"-13\"",    "Sleeve néoprène universelle","Pochette zip",    "Itel / Generic",        "Entrée",   49,  99, "Jumia ★"),
    # iPad Pro 11" M4 (2024)
    ("iPad Pro 11\" M4 (2024)","Generic Cover Pro 11 M4",    "Folio premium",   "Generic",               "Milieu",  280, 500, "Jumia"),
    ("iPad Pro 11\" M4 (2024)","ESR Rebound 360 Pro",        "Folio + clavier", "ESR",                   "Premium", 450, 750, "Amazon FR"),
    ("iPad Pro 11\" M4 (2024)","Apple Smart Folio Pro 11 M4","Smart Folio",     "Apple",                 "Premium", 950, 1300, "Apple"),
    ("iPad Pro 11\" M4 (2024)","Apple Magic Keyboard Pro 11","Folio + clavier", "Apple",                 "Premium",3500, 4500, "Apple — niche"),
    # iPad Air 13" M2 (2024)
    ("iPad Air 13\" M2 (2024)","Generic Folio 13\"",         "Folio grand fmt", "Generic",               "Milieu",  280, 450, "Jumia"),
    ("iPad Air 13\" M2 (2024)","Apple Smart Folio Air 13",   "Smart Folio",     "Apple",                 "Premium", 950, 1200, "Apple"),
    # iPad Mini 6 (2021)
    ("iPad Mini 6 (2021)",    "Generic Folio Mini 6",        "Folio compact",   "Generic / MoKo",        "Milieu",  120, 220, "Jumia/Amazon"),
    ("iPad Mini 6 (2021)",    "ESR Mini 6 Folio",            "Folio Pencil 2",  "ESR",                   "Milieu",  220, 380, "Amazon FR"),
    # iPad Air 5 (2022)
    ("iPad Air 5 (2022) 10.9\"","Generic Smart Folio Air 5", "Smart Folio",     "Generic",               "Milieu",  150, 280, "Jumia"),
    ("iPad Air 5 (2022) 10.9\"","ESR Rebound Pencil Holder", "Folio Pencil 2",  "ESR",                   "Milieu",  280, 450, "Amazon FR"),
    # iPad Pro 13" M4 (2024)
    ("iPad Pro 13\" M4 (2024)","Generic Cover Pro 13 M4",    "Folio grand fmt", "Generic",               "Milieu",  500, 850, "Jumia"),
    ("iPad Pro 13\" M4 (2024)","Apple Smart Folio Pro 13 M4","Smart Folio",     "Apple",                 "Premium",1100, 1400, "Apple"),
    # iPad Air M3
    ("iPad Air M3 11\" (2025)","Compatible Smart Folio M2",  "Smart Folio",     "Generic",               "Milieu",  180, 320, "Jumia (compat. M2)"),
    # iPad Air 4 (2020)
    ("iPad Air 4 (2020) 10.9\"","Generic Pochette Cover",    "Folio basique",   "Generic",               "Entrée",  100, 199, "Jumia"),
    # iPad Pro 11" 2018-2022
    ("iPad Pro 11\" (2018-2022)","Folio multi-gen Pro 11",   "Folio compatible","Generic",               "Milieu",  150, 300, "Jumia"),
    # iPad 7 / 8 (10.2")
    ("iPad 7 / 8 (10.2\")",   "Folio compatible 7/8/9",      "Folio multi-gen", "Itel / Generic",        "Entrée",   59, 130, "Jumia"),
    # iPad Pro 12.9" 2018-2022
    ("iPad Pro 12.9\" (2018-2022)","Folio multi-gen Pro 12.9","Folio grand fmt","Generic",               "Milieu",  200, 400, "Jumia"),
    # iPad Mini 5
    ("iPad Mini 5 (2019)",    "Folio Mini 7.9\"",            "Folio compact",   "Generic",               "Entrée",   60, 150, "Jumia/Avito"),
    # 2026 anticipation
    ("iPad Air M4 (2026)",    "Smart Folio compat. M2/M3",   "Smart Folio",     "Generic / ESR",         "Milieu",  200, 350, "Précommande"),
    ("iPad 12 (A19, 2026)",   "Folio Smart Cover (à venir)", "Folio basique",   "Generic",               "Entrée",  120, 230, "Précommande"),
    ("iPad Mini 8 OLED (2026)","Folio premium Mini OLED",    "Folio premium",   "Generic / Apple",       "Premium", 300, 600, "Précommande"),
    # P5 niche
    ("iPad Pro 9.7\" (2016)", "Folio sur commande",          "Folio basique",   "Generic",               "Entrée",   80, 180, "Sur commande"),
    ("iPad Air 2 (2014)",     "Folio sur commande",          "Folio basique",   "Generic",               "Entrée",   60, 150, "Sur commande"),
    ("iPad 5/6 (2017/2018)",  "Folio sur commande",          "Folio basique",   "Generic",               "Entrée",   60, 130, "Sur commande"),
]


def build_cases_catalog(ws: Worksheet) -> None:
    headers = ["iPad cible", "Référence pochette", "Type", "Marque",
               "Segment", "Prix min (MAD)", "Prix max (MAD)",
               "Prix moy. (MAD)", "Source / popularité"]
    write_title(ws, "Catalogue des pochettes par modèle iPad", 1, len(headers))

    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(headers))

    for i, row in enumerate(CASES_CATALOG, start=4):
        target, sku, ctype, brand, seg, p_min, p_max, src = row
        ws.cell(row=i, column=1, value=target)
        ws.cell(row=i, column=2, value=sku)
        ws.cell(row=i, column=3, value=ctype)
        ws.cell(row=i, column=4, value=brand)
        ws.cell(row=i, column=5, value=seg)
        ws.cell(row=i, column=6, value=p_min)
        ws.cell(row=i, column=7, value=p_max)
        ws.cell(row=i, column=8, value=f"=AVERAGE(F{i}:G{i})")
        ws.cell(row=i, column=9, value=src)

        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center",
                horizontal=("center" if c in (5, 6, 7, 8) else "left"),
                wrap_text=(c in (1, 2, 9)),
            )
            cell.font = Font(name="Calibri", size=10)
            if c in (6, 7, 8):
                cell.number_format = "#,##0 \"MAD\""
        ws.row_dimensions[i].height = 26

    last = 3 + len(CASES_CATALOG)
    seg_colors = {
        "Entrée":  GREEN_LIGHT,
        "Milieu":  BLUE_LIGHT,
        "Premium": AMBER_LIGHT,
    }
    for label, bg in seg_colors.items():
        rule = Rule(
            type="cellIs", operator="equal",
            formula=[f'"{label}"'],
            dxf=DifferentialStyle(fill=PatternFill("solid", start_color=bg, end_color=bg)),
        )
        ws.conditional_formatting.add(f"E4:E{last}", rule)

    autosize(ws, {"A": 30, "B": 36, "C": 22, "D": 22, "E": 12, "F": 14, "G": 14, "H": 14, "I": 24})
    freeze_header(ws, row=4)


# ────────────────────────────────────────────────────────────────────────
# Sheet 5: Sources
# ────────────────────────────────────────────────────────────────────────
SOURCES = [
    ("Jumia Maroc — Pochettes iPad",            "https://www.jumia.ma/mlp-pochette-ipad/"),
    ("Jumia Maroc — Étuis iPad",                "https://www.jumia.ma/mlp-etui-ipad/"),
    ("Jumia Maroc — Étuis & Protections",       "https://www.jumia.ma/etui-et-protection/"),
    ("Jumia Maroc — Pochette iPad 10",          "https://www.jumia.ma/mlp-pochette-ipad-10/"),
    ("Bestmark Maroc — Accessoires iPad",       "https://www.bestmark.ma/tous-nos-produits/accessoires/apple/accessoire-ipad.html"),
    ("Marjane Mall — Accessoires tablettes",    "https://www.marjanemall.ma/informatique-gaming/tablette/accessoires-tablettes"),
    ("Avito — Pochettes iPad d'occasion",       "https://www.avito.ma/fr/maroc/pochette_ipad--%C3%A0_vendre"),
    ("Apple — Comparatif iPad",                 "https://www.apple.com/ipad/compare/"),
    ("Macworld — iPad 2026 (A19)",              "https://www.macworld.com/article/3030272/2026-ipad-a19-design-display-specs-release-price.html"),
    ("Macworld — iPad Mini 2026 OLED",          "https://www.macworld.com/article/3060022/2026-ipad-mini-design-display-features-specs-release-date.html"),
    ("Macworld — iPad 2026 preview",            "https://www.macworld.com/article/3018973/2026-ipad-preview-what-to-expect-from-apples-next-tablet-lineup.html"),
    ("MacRumors — Roadmap iPad 2026",           "https://www.macrumors.com/2026/04/17/the-macrumors-show-whats-next-for-the-ipad/"),
    ("MacRumors — iPad Mini 8 OLED",            "https://www.macrumors.com/2026/04/15/oled-ipad-mini-release-pricing-what-to-expect/"),
    ("9to5Mac — iPads à venir 2026",            "https://9to5mac.com/2026/04/30/new-ipads-will-launch-later-this-year-heres-what-rumors-say-is-coming/"),
    ("Engadget — Best iPad cases 2026",         "https://www.engadget.com/computing/accessories/best-ipad-cases-to-protect-your-tablet-130033533.html"),
    ("Metapress — Best iPad case brands 2026",  "https://metapress.com/best-ipad-case-brands-in-2026-everyday-protection-smart-design-and-practical-features/"),
    ("Alibaba — iPad cases best buy 2025",      "https://www.alibaba.com/blog/bestselling-onsite/ipad-cases-best-buy-top-2025-picks-for-every-user.html"),
    ("Amazon — Best iPad cases (US)",           "https://www.amazon.com/best-ipad-case/s?k=best+ipad+case"),
]


def build_sources(ws: Worksheet) -> None:
    headers = ["#", "Source", "URL"]
    write_title(ws, "Sources consultées", 1, len(headers))

    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(headers))

    for i, (label, url) in enumerate(SOURCES, start=4):
        ws.cell(row=i, column=1, value=i - 3)
        ws.cell(row=i, column=2, value=label)
        link = ws.cell(row=i, column=3, value=url)
        link.hyperlink = url
        link.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        for c in range(1, 4):
            cell = ws.cell(row=i, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="center",
                horizontal=("center" if c == 1 else "left"),
                wrap_text=(c == 2),
            )
        ws.row_dimensions[i].height = 22

    autosize(ws, {"A": 6, "B": 50, "C": 80})
    freeze_header(ws, row=4)


# ────────────────────────────────────────────────────────────────────────
def main() -> None:
    wb = Workbook()

    # Default sheet → README
    build_readme(wb.active)

    build_top_stocking(wb.create_sheet("Top Stocking"))
    build_ipad_models(wb.create_sheet("iPad Models"))
    build_cases_catalog(wb.create_sheet("Cases Catalog"))
    build_sources(wb.create_sheet("Sources"))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"OK -> {OUT}  ({OUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
