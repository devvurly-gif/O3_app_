from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = Workbook()
sheet = wb.active
sheet.title = "Devis"

sheet.column_dimensions['A'].width = 40
sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 16
sheet.column_dimensions['D'].width = 16

header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)
section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
section_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
total_font = Font(bold=True, size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

sheet['A1'] = 'DEVIS - AMENAGEMENT CUISINE COMPLET'
sheet['A1'].font = Font(bold=True, size=16, color='1F4E78')
sheet.merge_cells('A1:D1')

sheet['A2'] = 'Fournitures, revetements et travaux complets'
sheet['A2'].font = Font(size=11, italic=True)
sheet.merge_cells('A2:D2')

row = 4
sheet[f'A{row}'] = 'Numero devis:'
sheet[f'B{row}'] = 'DV-2026-002'
row += 1
sheet[f'A{row}'] = 'Date:'
sheet[f'B{row}'] = datetime.now().strftime('%d/%m/%Y')
row += 1
sheet[f'A{row}'] = 'Budget total:'
sheet[f'B{row}'] = 200000
sheet[f'B{row}'].number_format = '#,##0 DH'
row += 1
sheet[f'A{row}'] = 'Validite:'
sheet[f'B{row}'] = '30 jours'
row += 2

headers = ['Designation', 'Quantite', 'Prix unitaire (DH)', 'Montant (DH)']
for col, header in enumerate(headers, 1):
    cell = sheet.cell(row=row, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

row += 1

# Section 1
sheet[f'A{row}'] = '1. FOURNITURES - MEUBLES ET EQUIPEMENTS'
sheet[f'A{row}'].font = section_font
sheet[f'A{row}'].fill = section_fill
for col in range(1, 5):
    sheet.cell(row=row, column=col).fill = section_fill
row += 1

fournitures = [
    ('Meuble base cuisine (L=240cm)', 1, 9500),
    ('Meuble haut cuisine (L=240cm)', 1, 6800),
    ('Plan de travail granit (m2)', 5, 1200),
    ('Credence verre laque (m2)', 4, 450),
    ('Evier inox 2 bacs + mitigeur premium', 1, 3200),
    ('Plaque cuisson 5 foyers gaz', 1, 4500),
    ('Four electrique integrable', 1, 5800),
    ('Hotte aspirante design (L=90cm)', 1, 3500),
    ('Lave-vaisselle integrable', 1, 5200),
    ('Micro-ondes integrable', 1, 2800),
    ('Poignees design et accessoires', 1, 1500),
]

fournitures_start = row
for designation, qty, price in fournitures:
    sheet[f'A{row}'] = designation
    sheet[f'B{row}'] = qty
    sheet[f'B{row}'].value = qty
    sheet[f'C{row}'] = float(price)
    sheet[f'C{row}'].value = float(price)
    sheet[f'D{row}'] = f'=B{row}*C{row}'

    sheet[f'B{row}'].alignment = Alignment(horizontal='center')
    sheet[f'B{row}'].number_format = '0'
    sheet[f'C{row}'].number_format = '#,##0'
    sheet[f'D{row}'].number_format = '#,##0'

    for col in range(1, 5):
        sheet.cell(row=row, column=col).border = border
    row += 1

fournitures_end = row - 1
fournitures_subtotal = row

sheet[f'A{row}'] = 'Sous-total Fournitures'
sheet[f'A{row}'].font = Font(bold=True, size=10, italic=True)
sheet[f'D{row}'] = f'=SUM(D{fournitures_start}:D{fournitures_end})'
sheet[f'D{row}'].font = Font(bold=True, size=10, italic=True)
sheet[f'D{row}'].number_format = '#,##0 DH'
row += 2

# Section 2
sheet[f'A{row}'] = '2. REVETEMENTS - SOL ET MURS'
sheet[f'A{row}'].font = section_font
sheet[f'A{row}'].fill = section_fill
for col in range(1, 5):
    sheet.cell(row=row, column=col).fill = section_fill
row += 1

revetements = [
    ('Carrelage sol porcelaine premium (m2)', 12, 550),
    ('Peinture murs cuisine (m2)', 35, 180),
    ('Plinthe carrelage (ml)', 15, 95),
    ('Fenetre avec double vitrage (m2)', 2, 3500),
    ('Porte interieure avec poignee', 1, 2200),
]

revetements_start = row
for designation, qty, price in revetements:
    sheet[f'A{row}'] = designation
    sheet[f'B{row}'] = qty
    sheet[f'C{row}'] = price
    sheet[f'D{row}'] = f'=B{row}*C{row}'

    sheet[f'B{row}'].alignment = Alignment(horizontal='center')
    sheet[f'C{row}'].number_format = '#,##0 DH'
    sheet[f'D{row}'].number_format = '#,##0 DH'

    for col in range(1, 5):
        sheet.cell(row=row, column=col).border = border
    row += 1

revetements_end = row - 1
revetements_subtotal = row

sheet[f'A{row}'] = 'Sous-total Revetements'
sheet[f'A{row}'].font = Font(bold=True, size=10, italic=True)
sheet[f'D{row}'] = f'=SUM(D{revetements_start}:D{revetements_end})'
sheet[f'D{row}'].font = Font(bold=True, size=10, italic=True)
sheet[f'D{row}'].number_format = '#,##0 DH'
row += 2

# Section 3
sheet[f'A{row}'] = '3. TRAVAUX D INSTALLATION'
sheet[f'A{row}'].font = section_font
sheet[f'A{row}'].fill = section_fill
for col in range(1, 5):
    sheet.cell(row=row, column=col).fill = section_fill
row += 1

travaux = [
    ('Demolition et evacuation cuisine', 1, 5000),
    ('Preparation et nettoyage chantier', 1, 2000),
    ('Pose meubles base et hauts', 2, 4500),
    ('Pose plan de travail et joints', 1, 3500),
    ('Pose carrelage sol (m2)', 12, 450),
    ('Pose credence (m2)', 4, 350),
    ('Pose plinthe et finitions sol', 1, 2200),
    ('Installation evier et mitigeur', 1, 2000),
    ('Installation plaque cuisson', 1, 1800),
    ('Installation four electrique', 1, 1800),
    ('Installation hotte et evacuation', 1, 2500),
    ('Installation lave-vaisselle', 1, 1500),
    ('Installation micro-ondes', 1, 800),
]

travaux_start = row
for designation, qty, price in travaux:
    sheet[f'A{row}'] = designation
    sheet[f'B{row}'] = qty
    sheet[f'C{row}'] = price
    sheet[f'D{row}'] = f'=B{row}*C{row}'

    sheet[f'B{row}'].alignment = Alignment(horizontal='center')
    sheet[f'C{row}'].number_format = '#,##0 DH'
    sheet[f'D{row}'].number_format = '#,##0 DH'

    for col in range(1, 5):
        sheet.cell(row=row, column=col).border = border
    row += 1

travaux_end = row - 1
travaux_subtotal = row

sheet[f'A{row}'] = 'Sous-total Travaux'
sheet[f'A{row}'].font = Font(bold=True, size=10, italic=True)
sheet[f'D{row}'] = f'=SUM(D{travaux_start}:D{travaux_end})'
sheet[f'D{row}'].font = Font(bold=True, size=10, italic=True)
sheet[f'D{row}'].number_format = '#,##0 DH'
row += 2

# Section 4
sheet[f'A{row}'] = '4. ELECTRICITE ET PLOMBERIE'
sheet[f'A{row}'].font = section_font
sheet[f'A{row}'].fill = section_fill
for col in range(1, 5):
    sheet.cell(row=row, column=col).fill = section_fill
row += 1

electro = [
    ('Renovation installation electrique', 1, 8500),
    ('Prises electriques et disjoncteurs', 8, 380),
    ('Eclairage LED integre meubles', 1, 3200),
    ('Eclairage plafond LED', 1, 2800),
    ('Interrupteurs et commandes', 1, 1500),
    ('Tuyauterie plomberie et raccords', 1, 4800),
    ('Robinetterie plomberie', 1, 2500),
    ('Evacuation eau installation', 1, 2200),
    ('Gaine evacuation hotte', 1, 1800),
]

electro_start = row
for designation, qty, price in electro:
    sheet[f'A{row}'] = designation
    sheet[f'B{row}'] = qty
    sheet[f'C{row}'] = price
    sheet[f'D{row}'] = f'=B{row}*C{row}'

    sheet[f'B{row}'].alignment = Alignment(horizontal='center')
    sheet[f'C{row}'].number_format = '#,##0 DH'
    sheet[f'D{row}'].number_format = '#,##0 DH'

    for col in range(1, 5):
        sheet.cell(row=row, column=col).border = border
    row += 1

electro_end = row - 1
electro_subtotal = row

sheet[f'A{row}'] = 'Sous-total Electricite et Plomberie'
sheet[f'A{row}'].font = Font(bold=True, size=10, italic=True)
sheet[f'D{row}'] = f'=SUM(D{electro_start}:D{electro_end})'
sheet[f'D{row}'].font = Font(bold=True, size=10, italic=True)
sheet[f'D{row}'].number_format = '#,##0 DH'
row += 2

# Section 5
sheet[f'A{row}'] = '5. FINITIONS ET SERVICES'
sheet[f'A{row}'].font = section_font
sheet[f'A{row}'].fill = section_fill
for col in range(1, 5):
    sheet.cell(row=row, column=col).fill = section_fill
row += 1

finitions = [
    ('Peinture murs et plafond', 1, 3500),
    ('Joints et scellement', 1, 2000),
    ('Nettoyage et evacuation debris', 1, 2500),
    ('Mise en service equipements', 1, 2000),
    ('Conseil et assistance', 1, 1500),
]

finitions_start = row
for designation, qty, price in finitions:
    sheet[f'A{row}'] = designation
    sheet[f'B{row}'] = qty
    sheet[f'C{row}'] = price
    sheet[f'D{row}'] = f'=B{row}*C{row}'

    sheet[f'B{row}'].alignment = Alignment(horizontal='center')
    sheet[f'C{row}'].number_format = '#,##0 DH'
    sheet[f'D{row}'].number_format = '#,##0 DH'

    for col in range(1, 5):
        sheet.cell(row=row, column=col).border = border
    row += 1

finitions_end = row - 1
finitions_subtotal = row

sheet[f'A{row}'] = 'Sous-total Finitions'
sheet[f'A{row}'].font = Font(bold=True, size=10, italic=True)
sheet[f'D{row}'] = f'=SUM(D{finitions_start}:D{finitions_end})'
sheet[f'D{row}'].font = Font(bold=True, size=10, italic=True)
sheet[f'D{row}'].number_format = '#,##0 DH'
row += 2

# Totaux
ht_row = row
sheet[f'A{row}'] = 'TOTAL HT (Hors Taxe)'
sheet[f'A{row}'].font = Font(bold=True, size=12)
sheet[f'D{row}'] = f'=D{fournitures_subtotal}+D{revetements_subtotal}+D{travaux_subtotal}+D{electro_subtotal}+D{finitions_subtotal}'
sheet[f'D{row}'].font = Font(bold=True, size=12)
sheet[f'D{row}'].number_format = '#,##0 DH'
sheet[f'D{row}'].border = border
row += 1

tva_row = row
sheet[f'A{row}'] = 'TVA 20%'
sheet[f'A{row}'].font = Font(bold=True, size=11)
sheet[f'D{row}'] = f'=D{ht_row}*0.20'
sheet[f'D{row}'].font = Font(bold=True, size=11)
sheet[f'D{row}'].number_format = '#,##0 DH'
sheet[f'D{row}'].border = border
row += 1

ttc_row = row
sheet[f'A{row}'] = 'TOTAL TTC'
sheet[f'A{row}'].font = total_font
sheet[f'A{row}'].fill = total_fill
sheet[f'D{row}'] = f'=D{ht_row}+D{tva_row}'
sheet[f'D{row}'].font = total_font
sheet[f'D{row}'].fill = total_fill
sheet[f'D{row}'].number_format = '#,##0 DH'
sheet[f'D{row}'].border = border

for col in range(1, 5):
    sheet.cell(row=row, column=col).fill = total_fill

wb.save('devis_cuisine_budget200k.xlsx')
print("Devis complet cree!")
